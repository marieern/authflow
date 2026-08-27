"""
app.py
======
Local web UI for the Fiori app -> authorization matrix workflow.

Run:
    pip install -r requirements.txt
    python app.py
Then open http://127.0.0.1:5000 in your browser.

This must run on a machine with network access to
fioriappslibrary.hana.ondemand.com — the sandbox this was built in has
none, so the live fetch path is untested end-to-end. See fiori_core.py's
docstring for what to verify first.
"""

import io
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import fiori_core

app = Flask(__name__)


@app.after_request
def disable_browser_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response

# In-memory index of everything fetched in the current server session,
# used for the shared-target-mapping cross-reference. Keyed by App ID.
SESSION_INDEX = {}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/fetch", methods=["POST"])
def api_fetch():
    payload = request.get_json(force=True)
    app_ids = []
    seen_app_ids = set()
    for value in payload.get("app_ids", []):
        if not isinstance(value, str):
            continue
        app_id = value.strip()
        if app_id and app_id.casefold() not in seen_app_ids:
            seen_app_ids.add(app_id.casefold())
            app_ids.append(app_id)
    refresh = bool(payload.get("refresh", False))
    templates = payload.get("templates", {})
    row_meta = payload.get("row_meta", {})  # {app_id: {module, function_group, persona}}
    existing_rows = {str(row.get("App ID", "")).casefold(): row
                     for row in payload.get("existing_rows", []) if row.get("App ID")}
    app_ids = [app_id for app_id in app_ids if app_id.casefold() not in existing_rows]
    if payload.get("require_business_user"):
        missing_business_users = [app_id for app_id in app_ids
                                  if not str(row_meta.get(app_id, {}).get("business_user", "")).strip()]
        if missing_business_users:
            return jsonify({
                "error": "Business User is required for new app(s): "
                         + ", ".join(missing_business_users)
            }), 400

    cache = fiori_core.load_cache()
    rows = []

    # First pass: fetch + stash semantic object/action for cross-referencing
    raw_by_app = {}
    for app_id in app_ids:
        raw_results, status = fiori_core.get_app_info(app_id, refresh=refresh, cache=cache)
        raw_by_app[app_id] = (raw_results, status)

        sem_obj = raw_results[0].get("SemanticObject") if raw_results else None
        sem_act = raw_results[0].get("SemanticAction") if raw_results else None
        SESSION_INDEX[app_id] = {"semantic_object": sem_obj, "semantic_action": sem_act}

    fiori_core.save_cache(cache)

    # Second pass: build enriched rows now that SESSION_INDEX is complete
    for app_id in app_ids:
        raw_results, status = raw_by_app[app_id]
        meta = row_meta.get(app_id, {})
        row = fiori_core.build_row(
            app_id=app_id,
            raw_results=raw_results,
            status=status,
            module=meta.get("module", ""),
            function_group=meta.get("function_group", ""),
            persona=meta.get("persona", ""),
            all_apps_index=SESSION_INDEX,
            templates=templates,
            end_users=meta.get("end_users", ""),
            business_user=meta.get("business_user", ""),
            pcg_role=meta.get("pcg_role", ""),
            derivation=meta.get("derivation", ""),
        )
        imported = existing_rows.get(app_id.casefold(), {})
        for field in (
            "End User(s)", "Business User", "Critical (confirmed)", "Function Group", "Persona",
            "Master Role (PFCG)", "Derived Role (PFCG)", "Custom Business Catalog", "Space", "Page",
        ):
            if imported.get(field) not in (None, ""):
                row[field] = imported[field]
        for field, value in imported.items():
            if field not in row and value not in (None, "") and field != "_raw":
                row[field] = value
        row["_raw"] = raw_results  # kept for the "raw fields" inspector panel
        rows.append(row)

    combined_rows = fiori_core.synchronize_shared_user_groups(
        list(existing_rows.values()) + rows, templates
    )
    updated_existing = {row["App ID"].casefold(): row for row in combined_rows
                        if row["App ID"].casefold() in existing_rows}
    rows = [row for row in combined_rows if row["App ID"].casefold() not in existing_rows]
    return jsonify({"rows": rows, "updated_existing": updated_existing})


@app.route("/api/import", methods=["POST"])
def api_import():
    upload = request.files.get("file")
    if not upload or not upload.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx or .xlsm Excel file."}), 400
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(upload, read_only=True, data_only=True)
        values = list(workbook.active.values)
        if not values:
            return jsonify({"error": "The workbook is empty."}), 400
        header_row_index = next(
            (index for index, row in enumerate(values[:30])
             if any(fiori_core.is_app_id_header(value) for value in row)),
            None,
        )
        if header_row_index is None:
            return jsonify({
                "error": "Could not find an App ID column in the first 30 rows. "
                         "Use a header such as App ID, Application ID, Fiori App, or App."
            }), 400
        headers = [str(value or "").strip() for value in values[header_row_index]]
        rows = fiori_core.normalize_matrix_rows(
            [dict(zip(headers, values_row)) for values_row in values[header_row_index + 1:] if any(values_row)]
        )
        rows = fiori_core.synchronize_shared_user_groups(rows)
        return jsonify({"rows": rows, "app_ids": [row["App ID"] for row in rows]})
    except (OSError, ValueError, TypeError) as error:
        return jsonify({"error": f"Could not parse the workbook: {error}"}), 400


@app.route("/api/export", methods=["POST"])
def api_export():
    payload = request.get_json(force=True)
    rows = payload.get("rows", [])

    columns = [
        "App ID", "App Title", "App Type (raw)", "Maintain/Display",
        "Semantic Object", "Semantic Action", "Target Mapping Note",
        "SAP Business Catalog", "SAP Business Catalog Descr",
        "SAP Role Template", "SAP Role Description",
        "Module", "Function Group", "Persona", "End User(s)", "Business User", "PCG Role (PFCG)",
        "Critical (suggested)", "Critical (confirmed)",
        "Master Role (PFCG)", "Derived Role (PFCG)", "Role Description (generated)",
        "Custom Business Catalog", "Custom Catalog Description",
        "Space", "Space Description", "Page", "Page Description",
        "Flags/Notes", "Status",
    ]
    for row in rows:
        for field in row:
            if field not in columns and field != "_raw":
                columns.append(field)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fiori App Matrix"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.font = body_font
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(38, len(col_name) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"fiori_app_matrix_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
